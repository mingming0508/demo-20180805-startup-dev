from openpyxl import load_workbook
import pytz
from urllib.request import urlopen, Request
from selenium import webdriver
import urllib
import bs4
import datetime

load_wb = load_workbook('./dialogflow/location.xlsx', data_only=True)
# load sheet name
sheet = load_wb['Sheet1']

# global x, y
axis = []

# api key
api_key = "8dxadbydcwiwucVyVbIke%2B0QRsADQ%2FKGW%2F9P76zE9NNZQFTQLVmc8otOVNJQeKUp%2FVceliBSvCxr2q5LsK1Z%2BA%3D%3D"

def crawling():
    # crawling
    url = 'https://fronteer.kr/service/kmaxy'
    # driver = webdriver.Chrome('C:\Users\minju\Anaconda3\chromedriver_win32')
    # req = request(url)
    # page = urlopen(req)
    # html = page.read()
    # soup = bs4.BeautifulSoup(html, 'html5lib')
    #driver.get(url)
    
    
# save excel data
allList = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet.min_column, max_col=5):
    a = []
    for cell in row:
        a.append(cell.value)
    allList.append(a)
print("done")


def get_api_date():
    standard_time = [2, 5, 8, 11, 14, 17, 20, 23]  # api response time
    time_now = datetime.datetime.now(tz=pytz.timezone('Asia/Seoul')).strftime('%H')  # get hour
    check_time = int(time_now) - 1
    day_calibrate = 0
    # hour to api time
    while not check_time in standard_time:
        check_time -= 1
        if check_time < 2:
            day_calibrate = 1  # yesterday
            check_time = 23

    date_now = datetime.datetime.now(tz=pytz.timezone('Asia/Seoul')).strftime('%Y%m%d')  # get date
    check_date = int(date_now) - day_calibrate

    # return date(yyyymmdd), tt00
    return str(check_date), (str(check_time) + '00')

def get_weather_data():
    api_date, api_time = get_api_date()
    url = "http://newsky2.kma.go.kr/service/SecndSrtpdFrcstInfoService2/ForecastSpaceData?"
    key = "serviceKey=" + api_key
    date = "&base_date=" + api_date
    time = "&base_time=" + api_time
    nx = f'&nx={axis[0]}'
    ny = f'&ny={axis[1]}'
    numOfRows = "&numOfRows=100"
    type_json = "&_type=json"
    api_url = url + key + date + time + nx + ny + numOfRows + type_json

    data = urlopen(api_url).read().decode('utf8')
    data_json = json.loads(data)

    print(data_json)

    parsed_json = data_json['response']['body']['items']['item']

    target_date = parsed_json[0]['fcstDate']  # get date and time
    target_time = parsed_json[0]['fcstTime']

    date_calibrate = target_date  # date of TMX, TMN
    if target_time > 1300:
        date_calibrate = str(int(target_date) + 1)

    passing_data = {}
    for one_parsed in parsed_json:
        if one_parsed['fcstDate'] == target_date and one_parsed['fcstTime'] == target_time:  # get today's data
            passing_data[one_parsed['category']] = one_parsed['fcstValue']

        if one_parsed['fcstDate'] == date_calibrate and (
                one_parsed['category'] == 'TMX' or one_parsed['category'] == 'TMN'):  # TMX, TMN at calibrated day
            passing_data[one_parsed['category']] = one_parsed['fcstValue']

    return passing_data

# bool
find = 0


# find x, y
def location(update, context, text):
    axis.clear()
    global find
    find = 0
    length = len(text)
    # length - 1 == allList
    for l in allList:
        if l[length - 1] == text[length - 1]:
            find = 1
            axis.append(l[3])
            axis.append(l[4])

    if find == 0:
        find_error(update, context)
    else:
        print(axis[0])
        print(axis[1])
        t = "If you want to know weather in there. Please enter /weather!"
        context.bot.send_message(chat_id=update.message.chat_id, text=t)
        
        
# result = get_weather_data()
#     t = f'The lowest degree is {result.get("TMN")}'
#     context.bot.send_message(chat_id=update.effective_chat.id, text=t)
#     t = f'The highest degree is {result.get("TMX")}'
#     context.bot.send_message(chat_id=update.effective_chat.id, text=t)
#     t = f'The humidity is {result.get("REH")}%'
#     context.bot.send_message(chat_id=update.effective_chat.id, text=t)
#     t = f'The rainfall probability is {result.get("POP")}%'
#     context.bot.send_message(chat_id=update.effective_chat.id, text=t)